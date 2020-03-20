import openpyxl, resourceDB, os

class userWallet:
    def __init__(self, test=None):
        self.base=resourceDB.resourceDataBase(user=1,test=test)
        self.wb=self.base.start()
        self.max_height=1001

    def PayInOut(self, amount):
        if(((type(amount)!=float) and (type(amount)!=int))): #zły typ danych
            raise Exception ("Wrong type of data")
        sheet=self.wb.active
        current_row=sheet.max_row
        if(current_row<2)and (amount<0): #wyplacenie gdy nie mamy srodkow bo niezainicjalizowane
            raise Exception ("There is no money")


        if(sheet.max_row>1):  #chyba nie potrzebne
            if (float(sheet.cell(column=1,row=current_row).value)<=0)and (amount<0):
                raise Exception("There is no money")

        if(sheet.max_row!=1):
            self.makeNewLine()
            current_row=sheet.max_row

            if sheet.cell(row=current_row, column=1).value+amount<0:
                amount+=sheet.cell(row=current_row, column=1).value
                sheet.cell(row=current_row, column=1).value=0
                self.wb.save(self.base.name)
                return amount

            sheet.cell(row=current_row, column=1).value+=amount
            
            #difference=sheet.cell(row=current_row-1, column=1).value-sheet.cell(row=current_row, column=1).value
            self.wb.save(self.base.name)
            return 0
        else:
            for i in range(len(self.base.attributes)):
                sheet.cell(column=i+1,row=2).value=0
            sheet.cell(column=1, row=2).value=amount
            self.wb.save(self.base.name)
            return amount

    def makeNewLine(self):
        sheet=self.wb.active
        if(sheet.max_row>self.max_height):
            i=2
            while i<self.max_height:
                for ii in range(len(self.base.attributes)):
                    sheet.cell(row=i,column=ii+1).value=sheet.cell(row=i+1,column=ii+1).value
                i+=1
        else:
            MAX_ROW=sheet.max_row+1
            for i in range(len(self.base.attributes)):
                sheet.cell(column=i+1, row=MAX_ROW).value=sheet.cell(column=i+1, row=MAX_ROW-1).value
        self.wb.save(self.base.name)

    def CurrentRow(self):
        lista=[0,0,0,0,0,0,0,0,0,0,0,0]
        if(self.wb.active.max_row==1):
            return lista
        sheet=self.wb.active
        current=sheet.max_row
        for i in range(len(self.base.attributes)):
            lista[i]=sheet.cell(column=i+1, row=current).value
        return lista
        
    def SellAndBuy(self, amount, attribute):
        #zwracana jest ilosc towaru ktory nie dal sie sprzedac 
        if(self.wb.active.max_row<3):
            raise Exception("There is no data in your wallet!")

        if(((type(amount)!=float) and (type(amount)!=int))): #zły typ danych
            raise Exception ("Wrong type of data")
        sheet=self.wb.active
        
        whichOne=0
        for i in range(len(self.base.attributes)):
            if(attribute==self.base.attributes[i]):
                whichOne=i
                break
        if (whichOne==0):
            raise Exception ("Wrong attribute")

        self.makeNewLine()
        current_row=sheet.max_row
        value=sheet.cell(row=current_row, column=whichOne+1).value+amount
        if(value>0):
           
            sheet.cell(row=current_row, column=whichOne+1).value=value
            
            self.wb.save(self.base.name)
            return 0 
        else:
           
            sheet.cell(row=current_row, column=whichOne+1).value=0
            self.wb.save(self.base.name)
            return value #zwracamy ile produktu nie udalo nam sie wymienic



def unitTest():
    if(os.path.exists("test.xlsx")):
        os.remove("test.xlsx")

    mistakes=0

    db=userWallet(test=-1)
    if(db.base.attributes!=["PLN","Oil","Gold","Copper","Silver","USD","EUR","CHF","Bitcoin","Etherum","Lisk","Litecoin"]):
        mistakes+=1

    if(db.wb.active.max_row!=1):
        mistakes+=1

    db.makeNewLine()
    if(db.wb.active.max_row!=2):
        mistakes+=1
    
    if(db.wb.active.cell(column=1,row=2).value!="PLN"):
        mistakes+=1

    db.wb.active.cell(column=1,row=2).value=33
    db.wb.active.cell(column=2,row=2).value=6
    db.wb.active.cell(column=3,row=2).value=10
    db.wb.active.cell(column=4,row=2).value=15
    db.wb.save(db.base.name)
    db.makeNewLine()
    if(db.wb.active.cell(column=1,row=3).value!=33):
        mistakes+=1

    if(db.wb.active.cell(column=2,row=3).value!=6):
        mistakes+=1

    if(db.wb.active.cell(column=3,row=3).value!=10):
        mistakes+=1

    if(db.wb.active.cell(column=4,row=3).value!=15):
        mistakes+=1

    if(os.path.exists("test.xlsx")):
        os.remove("test.xlsx")

    db=userWallet(test=-1)
    db.PayInOut(51)
    if(db.wb.active.cell(column=1,row=2).value!=51):
        mistakes+=1
    i=2
    while i<13:
        if(db.wb.active.cell(column=i,row=2).value!=0):
            mistakes+=1
        i+=1
    
    db.PayInOut(-52)
    if(db.wb.active.cell(column=1,row=3).value!=0):
        mistakes+=1
    i=2
    while i<13:
        if(db.wb.active.cell(column=i,row=3).value!=0):
            mistakes+=1
        i+=1
    
    mistakes+=1
    try:
        db.SellAndBuy(amount=14, attribute='a')
    except Exception as err:
        mistakes-=1
    
    row=db.wb.active.max_row
    db.SellAndBuy(amount=14, attribute='Copper')
    if(db.wb.active.max_row!=row+1):
        mistakes+=1
    mistakes+=1
    try:
        db.SellAndBuy(amount=14, attribute='PLN')
    except Exception as err:
        mistakes-=1
    

    
    lista=["PLN","Oil","Gold","Copper","Silver","USD","EUR","CHF","Bitcoin","Etherum","Lisk","Litecoin"]
    r=len(lista)-1
    for i in range(r):
        db.wb.active.cell(row=db.wb.active.max_row, column=i+2).value=0
    for i in lista:
        try:
            db.SellAndBuy(amount=17, attribute=i)
        except Exception as err:
            pass
    
    
    for i in range(r):
        if(db.wb.active.cell(row=db.wb.active.max_row,column=i+2).value!=17):
            mistakes+=1
    db.wb.save(db.base.name)
    
    for i in range(len(lista)+1):
        try:
            db.SellAndBuy(amount=-30, attribute=lista[i+1])
        except Exception as err:
            pass
        
    for i in range(r):
        if(db.wb.active.cell(row=db.wb.active.max_row,column=i+1).value!=0):
            mistakes+=1
    
    
    db.SellAndBuy(amount=100, attribute="Litecoin")
    db.wb.save(db.base.name)
    if(db.wb.active.cell(row=db.wb.active.max_row, column=db.wb.active.max_column).value!=117):
        mistakes+=1


    if(mistakes==0):
        print("unit test for userDB has been passed")
        return 1
    else:
        print("unit test for userDB has been failed")
        return 0


#unitTest()
    



#o=userWallet()
#i=o.SellAndBuy(attribute="Copper",amount=int(-30))
#print(i)

#o.PayInOut(int(39))

