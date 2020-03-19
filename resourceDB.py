import openpyxl, os, datetime

class resourceDataBase:
    def __init__(self, test=None, user=None):
        if  (test is None):
            if(user is None):
                self.name="Resources.xlsx"
                self.attributes=["Oil","Gold","Copper","Silver","USD","EUR","CHF","Bitcoin","Etherum","Lisk","Litecoin"]
            else:
                self.name="User.xlsx"
                self.attributes=["PLN","Oil","Gold","Copper","Silver","USD","EUR","CHF","Bitcoin","Etherum","Lisk","Litecoin"]
        else:
            self.name="test.xlsx"
            if(user is None):
                self.attributes=["Oil","Gold","Copper","Silver","USD","EUR","CHF","Bitcoin","Etherum","Lisk","Litecoin"]
            else:
                self.attributes=["PLN","Oil","Gold","Copper","Silver","USD","EUR","CHF","Bitcoin","Etherum","Lisk","Litecoin"]
        self.newComer=1

    def start(self):
        sheet=None
        if(not(os.path.exists(self.name))):
            sheet=self.createWorkbook()
        else:
            sheet=openpyxl.load_workbook(self.name)
            self.newComer=0

        return sheet 
    
    def isFresh(self):
        return self.newComer

    def createWorkbook(self):
        sheet=openpyxl.Workbook()
        current=sheet.active

        Font=openpyxl.styles.Font(name="title",bold=True,size=14)
        styl=openpyxl.styles.NamedStyle(name='title',font=Font)

        iterator=1
        for i in self.attributes:
            current.cell(row=1, column=iterator).value=i
            current.cell(row=1, column=iterator).style=styl
            iterator+=1
        sheet.save(self.name)
        return sheet


def unitTest():

    db=resourceDataBase(test=1)
    mistakes=0
    
    if(db.attributes!=["Oil","Gold","Copper","Silver","USD","EUR","CHF","Bitcoin","Etherum","Lisk","Litecoin"]):
        mistakes+=1
    if(db.newComer!=1):
        mistakes+=1
    if(db.isFresh()!=1):
        mistakes+=1
    if(db.name!="test.xlsx"):
        mistakes+=1

    if(os.path.exists("test.xlsx")):
        os.remove("test.xlsx")
    
    db.start()
    if(not(os.path.exists("test.xlsx"))):
        mistakes+=1
    wb=db.start()
    sheet=wb.active
    
    for i in range(len(db.attributes)):
        if(sheet.cell(row=1, column=i+1).value!=db.attributes[i]):
            mistakes+=1
    
    for i in range(len(db.attributes)):
        sheet.cell(row=2, column=i+1).value=i

    wb.save(db.name)
    db=''
    db=resourceDataBase(test=1)
    
    if(db.attributes!=["Oil","Gold","Copper","Silver","USD","EUR","CHF","Bitcoin","Etherum","Lisk","Litecoin"]):
        mistakes+=1
    db.start()
    if(db.newComer!=0):
        mistakes+=1
    if(db.isFresh()!=0):
        mistakes+=1

    if(db.name!="test.xlsx"):
        mistakes+=1

    wb=db.start()
    sheet=wb.active

    for i in range(len(db.attributes)):
        if(sheet.cell(row=2, column=i+1).value!=i):
            mistakes+=1
    

    if mistakes==0:
        print("unit test fo resourceDB has been passed")
        return 1
    else:
        print("unit test fo resourceDB has been failed")
        print(mistakes)
        return 0



    
    
    
#unitTest()

'''
o=resourceDataBase()
o.start()
'''
